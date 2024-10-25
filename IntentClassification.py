import openpyxl
import json

from langchain.chains.constitutional_ai.prompts import examples
from openpyxl import load_workbook
from langchain.chains import LLMChain
from langchain_community.chat_models import AzureChatOpenAI
from langchain.prompts import (
    ChatPromptTemplate,
    HumanMessagePromptTemplate,
    SystemMessagePromptTemplate
)

from ExampleSearchChromaDB import fetch_similar_queries

IntentClassificationPrompt = {
    "SYSTEM": """You are an intent identification bot. Based on the EMAIL_HISTORY, determine the Bot’s likely response and identify the intent of the bot's likely response.
     The identified intent should be selected from the list of INTENTS below.
        
        Prioritize email body over subject for intent identification.
        The EMAIL_HISTORY contains the conversation in chronological order, starting from the oldest to the most recent.
        If the intent of the bot’s likely response matches more than one intent, please provide the intent that most closely matches.
        Look into the EXAMPLES to identify the intent where applicable.

    INTENTS:
      1. ORDER_STATUS:
        - Handles emails inquiring about order status, ship dates, or delivery updates.

      2. PRODUCT_AVAILABILITY:
        - Handles emails inquiring about the availability of specific products, but not about product promotions or discounts.

      3. DAMAGES:
        - Handles emails reporting product damage, claims for damages, returning a damaged product or refused shipments. This includes any mention of damage in relation to a product.

      4. RETURNS:
        - Handles customer requests to return products, specifying that the products are not damaged, along with the reasons for returning the order. Replacements does not come under returns.

      5. TRADE_APPLICATION:
        - Handles inquiries and applications related to trade accounts, including requests to set up a new account or check the status of an existing one.

      6. BANTER: 
        - It handles inquiries that are usually involves general conversation, random comments, or all form of greetings are banter.

      7. OTHERS:
        - Any intent not captured above. This includes inquiries about return policies, general status updates, cancelling order, replacements, refund. Emails where the intent is unclear. If multiple intents are discussed or the email does not clearly inquire about a specific product or order, classify it as 'OTHERS'.

    """,
    "CONTEXT": """
    EMAIL_HISTORY:
    {email_history}
    
    EXAMPLES:
    {examples}
    """,
    "DISPLAY": """Ensure that the output is in the following JSON format exactly as shown:
        {{
          "intent": "[Main Intent Classified]"
        }}
        """,
    "REMEMBER": """Prioritize the email body for intent classification. classify it accordingly based on that context. Return the intent of bot likely response. Follow each intent description.""",
}


def gpt_call(GPT):
    gpt_config = {
        "4omini": {
            "OPENAI_API_KEY": "95058a9e99794e4689d179dd726e7eec",
            "OPENAI_DEPLOYMENT_NAME": "vassar-4o-mini",
            "OPENAI_EMBEDDING_MODEL_NAME": "vassar-text-ada002",
            "OPENAI_API_BASE": "https://vassar-openai.openai.azure.com/",
            "MODEL_NAME": "gpt-4o-mini",
            "OPENAI_API_TYPE": "azure",
            "OPENAI_API_VERSION": "2024-02-15-preview",
        }
    }

    config = gpt_config.get(GPT)
    if not config:
        print("Invalid GPT version specified")

    OPENAI_API_KEY = config["OPENAI_API_KEY"]
    OPENAI_DEPLOYMENT_NAME = config["OPENAI_DEPLOYMENT_NAME"]
    OPENAI_EMBEDDING_MODEL_NAME = config["OPENAI_EMBEDDING_MODEL_NAME"]
    OPENAI_API_BASE = config["OPENAI_API_BASE"]
    MODEL_NAME = config["MODEL_NAME"]
    OPENAI_API_TYPE = config["OPENAI_API_TYPE"]
    OPENAI_API_VERSION = config["OPENAI_API_VERSION"]

    llm = AzureChatOpenAI(
        deployment_name=OPENAI_DEPLOYMENT_NAME,
        model_name=MODEL_NAME,
        azure_endpoint=OPENAI_API_BASE,
        openai_api_type=OPENAI_API_TYPE,
        openai_api_key=OPENAI_API_KEY,
        openai_api_version=OPENAI_API_VERSION,
        temperature=0.0
    )
    return llm


def intent_classification(email_body, examples, GPT):
    # prompt = f"{IntentClassificationPrompt.get('SYSTEM_PROMPT')}\nEMAIL_HISTORY:\n{email_history}\nCURRENT_EMAIL:\n{user_latest_email}"
    llm = gpt_call(GPT)
    prompt_template = ChatPromptTemplate(
        messages=[
            SystemMessagePromptTemplate.from_template(IntentClassificationPrompt.get("SYSTEM")),
            HumanMessagePromptTemplate.from_template(IntentClassificationPrompt.get("CONTEXT")),
            SystemMessagePromptTemplate.from_template(IntentClassificationPrompt.get("DISPLAY")),
            SystemMessagePromptTemplate.from_template(IntentClassificationPrompt.get("REMEMBER"))
        ]
    )
    llm_chain = LLMChain(llm=llm, prompt=prompt_template, verbose=True)
    # print("prompt :: ", prompt_template)
    result = llm_chain.run({"email_history": email_body, "examples": examples})
    return result


file_path = '/home/saiprakesh/PycharmProjects/Prompt Testing Using Excel/automation_testing.xlsx'
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

for row in range(2, sheet.max_row + 1):
    if sheet[f'A{row}'].value is None:
        break
    cell_value = sheet[f'C{row}'].value
    if cell_value:
        email_history = cell_value.split("USER_LATEST_EMAIL:")[0].strip()
    else:
        email_history = ""
    expected_intent = sheet[f'D{row}'].value.strip()
    latest_body = sheet[f'C{row}'].value
    if latest_body:
        user_latest_email = sheet[f'C{row}'].value.split("USER_LATEST_EMAIL:")[1].strip()
    else:
        user_latest_email = ""
    examples = fetch_similar_queries(user_latest_email, top_k=10)
    email_body = email_history + user_latest_email
    classified_intent = intent_classification(email_body, examples, GPT="4omini")
    print(classified_intent)
    sheet[f'E{row}'] = json.loads(classified_intent).get("intent", "")
    if expected_intent.lower() == json.loads(classified_intent).get("intent", "").lower():
        sheet[f'F{row}'] = "PASS"
    else:
        sheet[f'F{row}'] = "FAIL"
    print(f"Intent Classification - {expected_intent == json.loads(classified_intent).get("intent", "")}")
    print("email_history: ", email_history)
    print("user_latest_email: ", user_latest_email)
    print("expected_intent: ", expected_intent)
    # print("examples: ", examples)

updated_file_path = '/home/saiprakesh/PycharmProjects/Prompt Testing Using Excel/automation_testing.xlsx'
workbook.save(updated_file_path)

print(f"Updated Excel file saved at: {updated_file_path}")
