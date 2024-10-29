import openpyxl
import json

from langchain.chains.constitutional_ai.prompts import examples
from openpyxl import load_workbook
from langchain.chains import LLMChain
from langchain_community.chat_models import AzureChatOpenAI
from langchain.prompts import (
    ChatPromptTemplate,
    SystemMessagePromptTemplate
)

from ExampleSearchChromaDB import fetch_similar_queries

IntentClassificationPrompt = """ You are an intent identification bot. Based on the CHAT_HISTORY and USER_QUERY, analyze the conversation and predict the likely response of the bot. Then, identify the intent of this likely response.

Steps:

1. Review the CHAT_HISTORY in chronological order (oldest message first, latest message last).
2. Consider the bot's potential response based on the history.
3. Match the likely response to one or more intents from the List of Possible Intents.
4. If the predicted response matches multiple intents, provide all matched intents along with their similarity scores.

Reference:
- EXAMPLES are available to help identify the appropriate intent where applicable.

Ensure that you classify the intent and return only the intent name.

Please choose the most appropriate intent from the following options:

1. Order Status: Handles messages specifically inquiring about the status of an order that has already been placed. This includes questions about shipping status, delivery updates, or order confirmation. It does not say anything about buying a product, which should be classified as "Product Availability."

2. Product Availability: Handles queries related to product availability, including if the user is asking about a product, expressing a desire to buy a product, or checking whether the product is in stock. If a user asks anything beyond availability, classify it as "Others." This intent also covers inquiries about the delivery timeline for products not yet ordered.

3. Returns: Handles customer requests explicitly related to returning a product which is not damaged. This intent only covers direct customer requests to return a product.

4. Damages: Handles queries related to damaged products. If a user mentions a damaged product, classify it as "Damages," even if they also request a return or refund for that product.

5. Trade Application: Only Handles queries related to inquiries or applications for trade accounts, credit terms, or setting up a business account, inquiring about loyalty levels. This can also include situations where a customer is inquiring about the status or continuity of a previously existing trade account.

6. Others: Handles all other queries, including queries related to refunds, canceled orders, modified orders, inquiries about promotions, or anything else outside of stock availability, order status, returns, damages, and trade applications. If the query is related to multiple intents,like asking for damages and order status in same query or something like that, then it should go to "Others."


Remember:
1. Output only the name of the intent.
2. Refer to previous conversation if you are not able to classify the intent from the user's question.
3. Always use the provided examples to assist in classifying the intent, along with the instructions above.
4. If the latest user query references refunds, cancellations, or order modifications, or contains multiple intents, classify it under "Others."

CHAT_HISTORY : {chat_history}
USER_QUERY : {question}

Below are examples to help clarify the process. Use them as a reference.
{examples}

"""


def gpt_call(GPT):
    gpt_config = {
        "4omini": {
            "OPENAI_API_KEY": "95058a9e99794e4689d179dd726e7eec",
            "OPENAI_DEPLOYMENT_NAME": "vassar-4o-mini",
            "OPENAI_EMBEDDING_MODEL_NAME": "vassar-text-ada002",
            "OPENAI_API_BASE": "https://vassar-openai.openai.azure.com/",
            "MODEL_NAME": "gpt-4o-mini",
            "OPENAI_API_TYPE": "azure",
            "OPENAI_API_VERSION": "2024-02-15-preview",
        }
    }

    config = gpt_config.get(GPT)
    if not config:
        print("Invalid GPT version specified")

    OPENAI_API_KEY = config["OPENAI_API_KEY"]
    OPENAI_DEPLOYMENT_NAME = config["OPENAI_DEPLOYMENT_NAME"]
    OPENAI_EMBEDDING_MODEL_NAME = config["OPENAI_EMBEDDING_MODEL_NAME"]
    OPENAI_API_BASE = config["OPENAI_API_BASE"]
    MODEL_NAME = config["MODEL_NAME"]
    OPENAI_API_TYPE = config["OPENAI_API_TYPE"]
    OPENAI_API_VERSION = config["OPENAI_API_VERSION"]

    llm = AzureChatOpenAI(
        deployment_name=OPENAI_DEPLOYMENT_NAME,
        model_name=MODEL_NAME,
        azure_endpoint=OPENAI_API_BASE,
        openai_api_type=OPENAI_API_TYPE,
        openai_api_key=OPENAI_API_KEY,
        openai_api_version=OPENAI_API_VERSION,
        temperature=0.0
    )
    return llm


def intent_classification(chat_history, user_latest_question_stripped, examples, GPT):
    llm = gpt_call(GPT)
    prompt_template = ChatPromptTemplate(
        messages=[
            SystemMessagePromptTemplate.from_template(IntentClassificationPrompt),
        ],
        input_variables=["CHAT_HISTORY", "question", "examples"],
    )
    llm_chain = LLMChain(llm=llm, prompt=prompt_template, verbose=False)
    result = llm_chain.run(
        {"chat_history": chat_history, "question": user_latest_question_stripped, "examples": examples})
    return result


file_path = '/home/saiprakesh/PycharmProjects/Prompt Testing Using Excel/automation_testing_chat.xlsx'
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active
count = 0
for row in range(2, sheet.max_row + 1):
    count = count + 1
    if sheet[f'A{row}'].value is None:
        break
    cell_value = sheet[f'C{row}'].value
    if cell_value:
        chat_history = cell_value.split("USER_LATEST_CHAT:")[0].strip()
    else:
        chat_history = ""
    expected_intent = sheet[f'D{row}'].value.strip()

    user_latest_question = sheet[f'C{row}'].value.split("USER_LATEST_CHAT:")[1].strip()

    _, _, user_latest_question_stripped = user_latest_question.partition("Human:")

    examples = fetch_similar_queries(user_latest_question_stripped, top_k=10)

    classified_intent = intent_classification(chat_history, user_latest_question_stripped, examples, GPT="4omini")
    print(count, ": ", classified_intent)
    if classified_intent == 'Banter':
        sheet[f'E{row}'] = 'OTHERS'
    else:
        sheet[f'E{row}'] = classified_intent
    if expected_intent.lower() == classified_intent.lower():
        sheet[f'F{row}'] = "PASS"
    else:
        sheet[f'F{row}'] = "FAIL"
    print(f"Intent Classification - {expected_intent.lower() == classified_intent.lower()}")
    print("email_history: ", chat_history)
    print("user_latest_email_body:", user_latest_question_stripped)
    print("expected_intent: ", expected_intent)
    print("examples: ", examples)

updated_file_path = '/home/saiprakesh/PycharmProjects/Prompt Testing Using Excel/automation_testing_chat.xlsx'
workbook.save(updated_file_path)

print(f"Updated Excel file saved at: {updated_file_path}")
