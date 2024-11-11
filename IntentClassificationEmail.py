from datetime import datetime
from venv import logger

import openpyxl
import json
import logging

from langchain.chains.constitutional_ai.prompts import examples
from openpyxl import load_workbook
from langchain.chains import LLMChain
from langchain_community.chat_models import AzureChatOpenAI
from langchain.prompts import (
    ChatPromptTemplate,
    HumanMessagePromptTemplate,
    SystemMessagePromptTemplate
)

from ExampleSearchChromaDB import fetch_similar_queries, fetch_similar_queries_for_intent
from IntentClassificationPromptEmail import IntentClassificationPrompt

logging.basicConfig(
    filename='intent_classification_with_reason_29_10_1.log',
    level=logging.WARNING,
    format='%(asctime)s - %(levelname)s - %(message)s'
)


class NoHTTPRequestsFilter(logging.Filter):
    def filter(self, record):
        return 'HTTP' not in record.getMessage()


def gpt_call(GPT):
    gpt_config = {
        "4omini": {
            "OPENAI_API_KEY": "95058a9e99794e4689d179dd726e7eec",
            "OPENAI_DEPLOYMENT_NAME": "vassar-4o-mini",
            "OPENAI_EMBEDDING_MODEL_NAME": "vassar-text-ada002",
            "OPENAI_API_BASE": "https://vassar-openai.openai.azure.com/",
            "MODEL_NAME": "gpt-4o-mini",
            "OPENAI_API_TYPE": "azure",
            "OPENAI_API_VERSION": "2024-02-15-preview",
        }
    }

    config = gpt_config.get(GPT)
    if not config:
        print("Invalid GPT version specified")

    OPENAI_API_KEY = config["OPENAI_API_KEY"]
    OPENAI_DEPLOYMENT_NAME = config["OPENAI_DEPLOYMENT_NAME"]
    OPENAI_EMBEDDING_MODEL_NAME = config["OPENAI_EMBEDDING_MODEL_NAME"]
    OPENAI_API_BASE = config["OPENAI_API_BASE"]
    MODEL_NAME = config["MODEL_NAME"]
    OPENAI_API_TYPE = config["OPENAI_API_TYPE"]
    OPENAI_API_VERSION = config["OPENAI_API_VERSION"]

    llm = AzureChatOpenAI(
        deployment_name=OPENAI_DEPLOYMENT_NAME,
        model_name=MODEL_NAME,
        azure_endpoint=OPENAI_API_BASE,
        openai_api_type=OPENAI_API_TYPE,
        openai_api_key=OPENAI_API_KEY,
        openai_api_version=OPENAI_API_VERSION,
        temperature=0.0
    )
    return llm


def intent_classification(email_body, examples, GPT):
    llm = gpt_call(GPT)
    prompt_template = ChatPromptTemplate(
        messages=[
            SystemMessagePromptTemplate.from_template(IntentClassificationPrompt.get("SYSTEM")),
            HumanMessagePromptTemplate.from_template(IntentClassificationPrompt.get("CONTEXT")),
            SystemMessagePromptTemplate.from_template(IntentClassificationPrompt.get("DISPLAY")),
            SystemMessagePromptTemplate.from_template(IntentClassificationPrompt.get("REMEMBER"))
        ]
    )
    llm_chain = LLMChain(llm=llm, prompt=prompt_template, verbose=False)
    result = llm_chain.run({"email_history": email_body, "examples": examples})
    return result


def sub_intent_classification(classified_intent, email_body, GPT):
    llm = gpt_call(GPT)
    _, _, user_latest_email = email_body.partition("\n")
    user_latest_email_body = user_latest_email.split("Body:")[1]
    examples = fetch_similar_queries_for_intent(user_latest_email_body, classified_intent, top_k=10)


file_path = '/home/saiprakesh/PycharmProjects/Prompt Testing Using Excel/automation_testing_email.xlsx'
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active
count = 0
formatted_time1 = ''
for row in range(2, sheet.max_row + 1):
    count = count + 1
    if count == 1:
        current_time1 = datetime.now()
        formatted_time1 = current_time1.strftime("%H:%M:%S")
    if sheet[f'A{row}'].value is None:
        break
    cell_value = sheet[f'C{row}'].value
    if cell_value:
        email_history = cell_value.split("USER_LATEST_EMAIL:")[0].strip()
    else:
        email_history = ""
    expected_intent = sheet[f'D{row}'].value.strip()
    latest_body = sheet[f'C{row}'].value
    if latest_body:
        user_latest_email = sheet[f'C{row}'].value.split("USER_LATEST_EMAIL:")[1].strip()
    else:
        user_latest_email = ""
    if len(user_latest_email.split("Body:")) > 1:
        user_latest_email_body = user_latest_email.split("Body:")[1]
    else:
        user_latest_email_body = user_latest_email.split("User:")[1]
    examples = fetch_similar_queries(user_latest_email_body, top_k=10)
    email_body = email_history + "\n" + user_latest_email
    classified_intent = intent_classification(email_body, examples, GPT="4omini")
    classified_sub_intent = ""
    if classified_intent.lower() == "ORDER_STATUS":
        classified_sub_intent = sub_intent_classification(classified_intent, email_body, GPT="4omini")
    print("classified_sub_intent : ", classified_sub_intent)
    print(count, ": ", classified_intent)
    sheet[f'E{row}'] = json.loads(classified_intent).get("intent", "")
    if expected_intent.lower() == json.loads(classified_intent).get("intent", "").lower():
        sheet[f'F{row}'] = "PASS"
    else:
        sheet[f'F{row}'] = "FAIL"
        logging.getLogger().addFilter(NoHTTPRequestsFilter())
        logging.error(
            "Failed Case #%d :: Expected intent :: '%s', but got  :: '%s'.\nEmail History:: %s\nUser Latest Email :: %s\nExamples:: %s\n",
            count, expected_intent, classified_intent, email_history, user_latest_email_body, examples
        )
        logging.error(
            "bot_likely_response ::  '%s', \n last_message ::  '%s', \n reason ::  '%s' \n",
            json.loads(classified_intent).get("bot_likely_response", ""),
            json.loads(classified_intent).get("last_message", ""),
            json.loads(classified_intent).get("reason", "")
        )
        print(f"Intent Classification - {expected_intent == json.loads(classified_intent).get("intent", "")}")
        print("email_history: ", email_history)
        print("user_latest_email: ", user_latest_email)
        print("user_latest_email_body:", user_latest_email_body)
        print("examples: ", examples)
        print("classified_intent:", json.loads(classified_intent).get("intent", ""))
        print("expected_intent: ", expected_intent)

updated_file_path = '/home/saiprakesh/PycharmProjects/Prompt Testing Using Excel/automation_testing_email.xlsx'
workbook.save(updated_file_path)
current_time2 = datetime.now()
formatted_time2 = current_time2.strftime("%H:%M:%S")
print("Start Time", formatted_time1)
print("END Time", formatted_time2)
print(f"Updated Excel file saved at: {updated_file_path}")

